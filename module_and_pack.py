from appJar import gui
import math
import openpyxl

app = gui("Geometry Calculator", "400x300")
datalist = []

def save_to_xls(results):
    for key, value in results.items():
        datalist.append({key: value})
    wb = openpyxl.Workbook()
    ws = wb.active
    for idx, data in enumerate(datalist, start=1):
        for column, (key, value) in enumerate(data.items(), start=1):
            ws.cell(row=idx, column=1, value=key)
            ws.cell(row=idx, column=2, value=value)
    wb.save("results.xlsx")

def calculate_rectangle(btn):
    length = int(app.getEntry("Rectangle Length"))
    width = int(app.getEntry("Rectangle Width"))
    area = length * width
    app.setLabel("Rectangle Area", "Area: " + str(area))
    save_to_xls({"Rectangle Area": area})

def calculate_triangle(btn):
    base = int(app.getEntry("Triangle Base"))
    height = int(app.getEntry("Triangle Height"))
    area = 0.5 * base * height
    app.setLabel("Triangle Area", "Area: " + str(area))
    save_to_xls({"Triangle Area": area})

def calculate_trapezoid(btn):
    base1 = int(app.getEntry("Trapezoid Base 1"))
    base2 = int(app.getEntry("Trapezoid Base 2"))
    height = int(app.getEntry("Trapezoid Height"))
    area = 0.5 * (base1 + base2) * height
    app.setLabel("Trapezoid Area", "Area: " + str(area))
    save_to_xls({"Trapezoid Area": area})

def calculate_circles(btn):
    radius = float(app.getEntry("Circle Radius"))
    area = math.pi * radius**2
    circumradius = 2 * radius
    inscribedradius = radius / (math.sqrt(2))
    app.setLabel("Circle Area", "Area: " + str(area))
    app.setLabel("Circumcircle Radius", "Circumcircle Radius: " + str(circumradius))
    app.setLabel("Inscribed Circle Radius", "Inscribed Circle Radius: " + str(inscribedradius))
    save_to_xls({"Inscribed Circle Radius": inscribedradius})

app.addLabelEntry("Rectangle Length")
app.addLabelEntry("Rectangle Width")
app.addButton("Calculate Rectangle Area", calculate_rectangle)
app.addLabel("Rectangle Area", "")

app.addLabelEntry("Triangle Base")
app.addLabelEntry("Triangle Height")
app.addButton("Calculate Triangle Area", calculate_triangle)
app.addLabel("Triangle Area", "")

app.addLabelEntry("Trapezoid Base 1")
app.addLabelEntry("Trapezoid Base 2")
app.addLabelEntry("Trapezoid Height")
app.addButton("Calculate Trapezoid Area", calculate_trapezoid)
app.addLabel("Trapezoid Area", "")

app.addLabelEntry("Circle Radius")
app.addButton("Calculate Circle Properties", calculate_circles)
app.addLabel("Circle Area", "")
app.addLabel("Circumcircle Radius", "")
app.addLabel("Inscribed Circle Radius", "")

app.go()
